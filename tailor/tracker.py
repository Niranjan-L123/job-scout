"""Append an application row to Job_Applications_Tracker.xlsx.

Columns (existing sheet 'Applications'):
  Company | Role | Location | Date Applied | Status | CV Used | Cover Letter | Job Link | Notes

Date Applied / Status are left blank on purpose — they're filled in when the
user actually submits. CV Used is stored as a path relative to the resume root.

    python -m tailor.tracker --company Monzo --role "Analytics Engineer Intern" \
        --location London --cv "monzo/..._Monzo.docx" --link <url> [--notes ...]
"""
import argparse

import openpyxl

from . import RESUME_ROOT, TRACKER

HEADERS = ["Company", "Role", "Location", "Date Applied", "Status",
           "CV Used", "Cover Letter", "Job Link", "Notes"]


def append_row(company, role, location="", cv="", link="", notes="",
               cover_letter="No", tracker_path=TRACKER):
    wb = openpyxl.load_workbook(tracker_path)
    ws = wb["Applications"] if "Applications" in wb.sheetnames else wb.active
    # Normalise a CV path to be relative to the resume root when possible.
    cv_rel = cv
    try:
        cv_rel = str((RESUME_ROOT / cv).relative_to(RESUME_ROOT)) if cv else ""
    except Exception:
        cv_rel = cv
    ws.append([company, role, location, None, None, cv_rel, cover_letter,
               link, notes or None])
    wb.save(tracker_path)
    return ws.max_row


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--company", required=True)
    ap.add_argument("--role", required=True)
    ap.add_argument("--location", default="")
    ap.add_argument("--cv", default="")
    ap.add_argument("--link", default="")
    ap.add_argument("--notes", default="")
    ap.add_argument("--tracker", default=str(TRACKER))
    args = ap.parse_args()
    row = append_row(args.company, args.role, args.location, args.cv,
                     args.link, args.notes, tracker_path=args.tracker)
    print(f"appended row {row}: {args.company} | {args.role}")


if __name__ == "__main__":
    main()
